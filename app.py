# app.py — Project Apex 2035 | Portfolio Dashboard
# Run: streamlit run app.py

import json
import datetime
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from config import (
    PROJECT_NAME, TARGET_5X_USD, TARGET_10X_USD,
    BROKERS, ACTIVE_BROKERS, LEGACY_BROKERS,
    REPORT_CURRENCIES, DEFAULT_CURRENCY,
)
from core.prices import get_prices_batch, get_hkd_usd_rate
from concurrent.futures import ThreadPoolExecutor
from core.engine import (
    build_portfolio, portfolio_summary, allocation_by,
    concentration_alerts, compliance_check,
    calc_new_avg_cost, target_progress,
    calculate_pillars, get_technical_signals,
)
from core.exports import (
    export_portfolio_pdf, export_stock_pdf, export_conviction_pdf,
)
from core.lseg_data import (
    lseg_available, lseg_connected, lseg_desktop_available, refresh_lseg,
)

# ── Page config ───────────────────────────────────────────────────
st.set_page_config(
    page_title="Apex 2035",
    page_icon="📈",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ── Custom CSS ────────────────────────────────────────────────────
st.markdown("""
<style>
    .block-container { padding-top: 1rem; }
    .metric-card {
        background: #1a1a2e; border-radius: 8px;
        padding: 14px 18px; margin-bottom: 8px;
    }
    .pos { color: #4ade80; font-weight: 600; }
    .neg { color: #f87171; font-weight: 600; }
    .neu { color: #94a3b8; }
    .target-bar { background: #0f172a; border-radius: 6px;
                  padding: 10px 14px; margin: 4px 0; }
    div[data-testid="stTabs"] button { font-size: 14px; font-weight: 500; }
    .stDataFrame { font-size: 12px; }
</style>
""", unsafe_allow_html=True)

# ── Determine if Google Sheets is configured ───────────────────────
SHEETS_AVAILABLE = "gcp_service_account" in st.secrets

if SHEETS_AVAILABLE:
    from core.sheets import (
        read_holdings, append_trades, read_trades,
        append_portfolio_snapshot, read_portfolio_history,
    )

    @st.cache_resource
    def init_connections():
        """Pre-warm Google Sheets connection on startup so first tab load is faster."""
        try:
            from core.sheets import get_client
            get_client()
            return True
        except Exception:
            return False

    init_connections()
else:
    # Offline mode: load from config.py directly (no Google Sheets needed for demo)
    from config import INITIAL_POSITIONS, HKD_USD_RATE
    import core.engine as engine

    def _build_offline_holdings() -> pd.DataFrame:
        rows = []
        for pos in INITIAL_POSITIONS:
            brokers = pos.get("brokers", [])
            total_shares = sum(b["shares"] for b in brokers)
            total_cost_w = sum(b["shares"] * b["avg_cost_local"] for b in brokers)
            avg_cost_l   = total_cost_w / total_shares if total_shares else 0
            avg_cost_usd = avg_cost_l / HKD_USD_RATE if pos["ccy"] == "HKD" else avg_cost_l
            rows.append({
                "ticker": pos["ticker"], "name": pos["name"],
                "region": pos["region"], "sector": pos["sector"],
                "barbell_class": pos["barbell_class"], "ccy": pos["ccy"],
                "total_shares": total_shares,
                "avg_cost_local": avg_cost_l, "avg_cost_usd": avg_cost_usd,
                "brokers_json": json.dumps(pos.get("brokers", [])),
                "manual_price": None, "compliance_flag": "",
                "lockup_expiry": "", "notes": "", "last_updated": "",
            })
        return pd.DataFrame(rows)

    def read_holdings():
        return _build_offline_holdings()

    def append_trades(trades, source="manual"):
        st.warning("Google Sheets not configured — trades not persisted. See SETUP.md.")


# ─────────────────────────────────────────────────────────────────
# SIDEBAR — global controls
# ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("⚙️ Controls")
    report_ccy = st.radio("Report currency", REPORT_CURRENCIES,
                           index=REPORT_CURRENCIES.index(DEFAULT_CURRENCY),
                           horizontal=True)
    ccy_sym = "HK$" if report_ccy == "HKD" else "$"

    st.divider()
    if st.button("🔄 Refresh prices"):
        st.cache_data.clear()
        st.rerun()

    st.divider()
    # Technical signals overlay (Master Ledger tab)
    show_ta = st.checkbox(
        "📊 Show technical signals",
        value=False,
        help="Adds 200DMA, 52W Range, RS, Volume columns to Master Ledger. "
             "Fetches data for visible positions — adds ~5-10 s.",
    )

    st.divider()
    # LSEG status — desktop session only, local PC only
    _lseg_ok = lseg_desktop_available()
    if _lseg_ok:
        st.success("🔬 LSEG available")
        st.caption("Refinitiv Workspace detected")
        _lseg_n = st.session_state.get("lseg_calls", 0)
        if _lseg_n:
            st.caption(f"LSEG calls this session: {_lseg_n}")
        if st.button("↺ Reconnect LSEG", key="lseg_reconnect"):
            refresh_lseg()
            st.rerun()
    else:
        st.info("📊 yfinance mode")
        st.caption(
            "Open Refinitiv Workspace on this PC "
            "to enable LSEG data"
        )

    st.divider()
    st.caption(f"Project Apex 2035\nTarget: {ccy_sym}{TARGET_5X_USD:,.0f}\nHK tax: 0% CGT ✓")

    if not SHEETS_AVAILABLE:
        st.warning("⚠️ Offline mode\nGoogle Sheets not connected.\nSee SETUP.md to connect.")


# ─────────────────────────────────────────────────────────────────
# DATA LOAD
# ─────────────────────────────────────────────────────────────────
@st.cache_data(ttl=300)
def load_data(report_ccy: str):
    holdings_df = read_holdings()
    tickers = tuple(holdings_df["ticker"].tolist())
    prices  = get_prices_batch(tickers)
    fx_rate = get_hkd_usd_rate()
    port_df = build_portfolio(holdings_df, prices, fx_rate, report_ccy)
    summary = portfolio_summary(port_df, report_ccy)
    progress = target_progress(summary["total_mv"] if report_ccy == "USD"
                               else summary["total_mv"] / fx_rate)
    # Auto-save daily snapshot (silently — never crashes app)
    if SHEETS_AVAILABLE:
        try:
            mv_usd = summary["total_mv"] if report_ccy == "USD" else summary["total_mv"] / fx_rate
            append_portfolio_snapshot({
                "total_mv_usd":   mv_usd,
                "total_cost_usd": summary["total_cost"] if report_ccy == "USD"
                                  else summary["total_cost"] / fx_rate,
                "total_gl_usd":   summary["total_gl"] if report_ccy == "USD"
                                  else summary["total_gl"] / fx_rate,
                "gl_pct":         summary["total_gl_pct"],
                "hkd_usd_rate":   fx_rate,
            })
        except Exception:
            pass
    return port_df, summary, progress, fx_rate


with st.spinner("Loading portfolio…"):
    port_df, summary, progress, fx_rate = load_data(report_ccy)

total_mv   = summary["total_mv"]
total_gl   = summary["total_gl"]
total_cost = summary["total_cost"]


# ─────────────────────────────────────────────────────────────────
# HEADER
# ─────────────────────────────────────────────────────────────────
col_title, col_ts = st.columns([3, 1])
with col_title:
    st.title(f"📈 {PROJECT_NAME}")
with col_ts:
    st.caption(f"FX: USD/HKD = {fx_rate:.4f}\nPrices ~15 min delayed")

# Top KPI row
k1, k2, k3, k4, k5, k6 = st.columns(6)
gl_color = "normal" if total_gl >= 0 else "inverse"
k1.metric("Total Portfolio",
          f"{ccy_sym}{total_mv/1e6:.3f}M",
          f"{ccy_sym}{total_gl/1e6:+.3f}M")
k2.metric("Unrealized G/L",
          f"{ccy_sym}{total_gl:,.0f}",
          f"{total_gl/total_cost*100:+.1f}%" if total_cost else "—")
k3.metric("Positions", summary["n_positions"])
k4.metric("5x Target",
          f"{ccy_sym}{TARGET_5X_USD/1e6:.2f}M",
          f"{progress['pct_to_5x']:.1f}% there")
k5.metric("CAGR needed → 5x",
          f"{progress['cagr_needed_5x']:.1f}%",
          "by 2035")
k6.metric("Price source", "yfinance", "~15 min delay")

st.divider()

# ─────────────────────────────────────────────────────────────────
# TABS
# ─────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8, tab9 = st.tabs([
    "📊 Master Ledger",
    "🏦 Broker Recon",
    "🎯 Analytics",
    "⚠️ Alerts",
    "✏️ Trade Entry",
    "📄 Export",
    "📈 Stock Analyzer",
    "🎯 Conviction Tracker",
    "🌊 Flow Monitor",
])


# ══════════════════════════════════════════════════════════════════
# TAB 1 — MASTER LEDGER
# ══════════════════════════════════════════════════════════════════
with tab1:
    st.subheader("Consolidated Holdings — Master Ledger")

    # Filter controls
    fc1, fc2, fc3, fc4 = st.columns(4)
    with fc1:
        region_filter = st.multiselect("Region",
            sorted(port_df["region"].unique()), default=[])
    with fc2:
        sector_filter = st.multiselect("Sector",
            sorted(port_df["sector"].unique()), default=[])
    with fc3:
        barbell_filter = st.multiselect("Barbell",
            ["CORE", "TACTICAL", "SPECULATIVE"], default=[])
    with fc4:
        search = st.text_input("Search ticker / name", "")

    df_view = port_df.copy()
    if region_filter:  df_view = df_view[df_view["region"].isin(region_filter)]
    if sector_filter:  df_view = df_view[df_view["sector"].isin(sector_filter)]
    if barbell_filter: df_view = df_view[df_view["barbell_class"].isin(barbell_filter)]
    if search:
        s = search.lower()
        df_view = df_view[
            df_view["ticker"].str.lower().str.contains(s) |
            df_view["name"].str.lower().str.contains(s)
        ]

    df_view = df_view.sort_values("mv_usd", ascending=False, na_position="last")

    # Build display dataframe
    def fmt_price(row):
        if row["live_price"] is None:
            return "—"
        sym = "HK$" if row["ccy"] == "HKD" else "$"
        return f"{sym}{row['live_price']:,.2f}"

    def fmt_mv(val):
        if pd.isna(val): return "—"
        return f"{ccy_sym}{val:,.0f}"

    def fmt_gl(val, pct):
        if pd.isna(val): return "—"
        sign = "+" if val >= 0 else ""
        return f"{sign}{ccy_sym}{val:,.0f} ({sign}{pct:.1f}%)" if pd.notna(pct) else f"{sign}{ccy_sym}{val:,.0f}"

    display = pd.DataFrame({
        "Ticker":       df_view["ticker"],
        "Name":         df_view["name"],
        "Region":       df_view["region"],
        "Sector":       df_view["sector"],
        "Barbell":      df_view["barbell_class"],
        "Shares":       df_view["shares"].apply(lambda x: f"{x:,.4f}".rstrip('0').rstrip('.') if x % 1 != 0 else f"{x:,.0f}"),
        "Price (Local)":df_view.apply(fmt_price, axis=1),
        f"MV ({report_ccy})": df_view["mv_report"].apply(fmt_mv),
        f"Cost ({report_ccy})":df_view["cost_usd"].apply(
            lambda x: fmt_mv(x * fx_rate) if report_ccy == "HKD" else fmt_mv(x)),
        "G/L":          df_view.apply(lambda r: fmt_gl(r["gl_report"], r["gl_pct"]), axis=1),
        "Held at":      df_view["brokers"],
        "Price time":   df_view["price_ts"],
    })

    # ── Technical Signals overlay ─────────────────────────────────
    if show_ta:
        _vis_tickers = df_view["ticker"].tolist()
        with st.spinner(f"Calculating technical signals for {len(_vis_tickers)} positions…"):
            with ThreadPoolExecutor(max_workers=3) as _ex:
                _futures = {t: _ex.submit(get_technical_signals, t) for t in _vis_tickers}
            _ta_map = {t: f.result() for t, f in _futures.items()}

        def _fmt_ma(ticker):
            s = _ta_map.get(ticker, {}).get("ma200", {})
            if not s:
                return "—"
            pct = s["pct_from_ma"]
            return f"✅ +{pct:.1f}%" if s["above"] else f"❌ {pct:.1f}%"

        def _fmt_52w(ticker):
            s = _ta_map.get(ticker, {}).get("range_52w", {})
            if not s:
                return "—"
            p = s["position_pct"]
            if p > 75:
                return f"🟢 {p:.0f}%"
            elif p >= 25:
                return f"🟡 {p:.0f}%"
            else:
                return f"🔴 {p:.0f}%"

        def _fmt_rs(ticker):
            s = _ta_map.get(ticker, {}).get("rs_vs_index", {})
            if not s:
                return "—"
            rel = s["relative_pct"]
            bm  = s.get("benchmark", "SPY")
            sign = "+" if rel >= 0 else ""
            return f"{sign}{rel:.1f}% vs {bm}"

        def _fmt_vol(ticker):
            s = _ta_map.get(ticker, {}).get("volume", {})
            if not s:
                return "—"
            r = s["ratio"]
            if r > 1.5:
                return f"🔥 {r:.1f}x"
            elif r >= 0.75:
                return f"📊 {r:.1f}x"
            else:
                return f"😴 {r:.1f}x"

        display["200DMA"]   = display["Ticker"].apply(_fmt_ma)
        display["52W Range"] = display["Ticker"].apply(_fmt_52w)
        display["RS 3M"]    = display["Ticker"].apply(_fmt_rs)
        display["Volume"]   = display["Ticker"].apply(_fmt_vol)

    st.dataframe(display, use_container_width=True, hide_index=True,
                 height=min(50 + len(display) * 35, 650))

    # Summary footer
    valid = df_view[df_view["mv_report"].notna()]
    f1, f2, f3 = st.columns(3)
    f1.metric("Filtered MV", f"{ccy_sym}{valid['mv_report'].sum():,.0f}")
    f2.metric("Filtered Cost",
              f"{ccy_sym}{valid['cost_usd'].sum() * (fx_rate if report_ccy=='HKD' else 1):,.0f}")
    gl_filt = valid["gl_report"].sum() if "gl_report" in valid else 0
    f3.metric("Filtered G/L", f"{ccy_sym}{gl_filt:,.0f}")

    # PDF export
    st.divider()
    _pdf1_col, _ = st.columns([1, 3])
    with _pdf1_col:
        if st.button("📄 Generate Portfolio PDF", key="pdf_btn_tab1",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf1 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF",
                _pdf1,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab1",
            )


# ══════════════════════════════════════════════════════════════════
# TAB 2 — BROKER RECONCILIATION
# ══════════════════════════════════════════════════════════════════
with tab2:
    st.subheader("Broker Reconciliation — Cross-Check View")
    st.caption("Match these numbers against your broker mobile apps.")

    for broker in BROKERS:
        broker_rows = []
        for _, row in port_df.iterrows():
            broker_detail = next(
                (b for b in row["brokers_list"] if b["broker"] == broker), None
            )
            if broker_detail is None:
                continue

            shares_here = broker_detail["shares"]
            cost_here   = broker_detail["avg_cost_local"]
            live_price  = row["live_price"]
            ccy         = row["ccy"]
            sym         = "HK$" if ccy == "HKD" else "$"

            if live_price:
                mv_local_here = shares_here * live_price
                mv_usd_here   = mv_local_here / fx_rate if ccy == "HKD" else mv_local_here
                mv_report_here = mv_usd_here * fx_rate if report_ccy == "HKD" else mv_usd_here
                mv_str = f"{ccy_sym}{mv_report_here:,.0f}"
            else:
                mv_report_here = None
                mv_str = "—"

            broker_rows.append({
                "Ticker":          row["ticker"],
                "Name":            row["name"],
                "Shares":          f"{shares_here:,.4f}".rstrip('0').rstrip('.') if shares_here % 1 != 0 else f"{shares_here:,.0f}",
                "Avg Cost":        f"{sym}{cost_here:,.4f}",
                "Live Price":      f"{sym}{live_price:,.2f}" if live_price else "—",
                f"MV ({report_ccy})": mv_str,
            })

        total_broker_mv = sum(
            float(r[f"MV ({report_ccy})"].replace(ccy_sym, "").replace(",", ""))
            for r in broker_rows if r[f"MV ({report_ccy})"] != "—"
        ) if broker_rows else 0

        n = len(broker_rows)
        label = f"{'🟢' if n > 0 else '⚪'} {broker}  —  {n} position{'s' if n != 1 else ''}  |  est. {ccy_sym}{total_broker_mv:,.0f} {report_ccy}"
        with st.expander(label, expanded=(n > 0)):
            if broker_rows:
                st.dataframe(pd.DataFrame(broker_rows),
                             use_container_width=True, hide_index=True)
            else:
                st.caption("No positions recorded for this broker.")

    st.divider()
    _pdf2_col, _ = st.columns([1, 3])
    with _pdf2_col:
        if st.button("📄 Generate Portfolio PDF", key="pdf_btn_tab2",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf2 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF",
                _pdf2,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab2",
            )


# ══════════════════════════════════════════════════════════════════
# TAB 3 — ANALYTICS (Barbell + Sector + Region)
# ══════════════════════════════════════════════════════════════════
with tab3:
    st.subheader("Portfolio Analytics")

    view_mode = st.radio("Group by", ["Barbell Class", "Sector", "Region"],
                          horizontal=True)
    col_map = {"Barbell Class": "barbell_class", "Sector": "sector", "Region": "region"}
    group_col = col_map[view_mode]

    alloc = allocation_by(port_df, group_col)

    c_pie, c_bar = st.columns(2)
    with c_pie:
        COLOR_MAP = {
            # Barbell
            "CORE":        "#2E75B6", "TACTICAL": "#1D9E75", "SPECULATIVE": "#D85A30",
            # Region
            "US":          "#378ADD", "HK": "#1D9E75", "China": "#D85A30",
            "SEA":         "#D4537E", "Other": "#7F77DD",
        }
        colors = [COLOR_MAP.get(g, "#888") for g in alloc[group_col]]
        fig_pie = px.pie(alloc, values="mv_usd", names=group_col,
                         title=f"Allocation by {view_mode}",
                         color_discrete_sequence=colors)
        fig_pie.update_traces(textposition="inside", textinfo="percent+label")
        fig_pie.update_layout(showlegend=False, margin=dict(t=40, b=0, l=0, r=0))
        st.plotly_chart(fig_pie, use_container_width=True)

    with c_bar:
        fig_bar = px.bar(alloc.sort_values("mv_usd"),
                         x="mv_usd", y=group_col, orientation="h",
                         title=f"Market Value by {view_mode} ({report_ccy})",
                         color=group_col, color_discrete_map=COLOR_MAP,
                         text=alloc["pct"].apply(lambda x: f"{x:.1f}%"))
        fig_bar.update_traces(textposition="outside")
        fig_bar.update_layout(showlegend=False, xaxis_title=f"MV ({report_ccy})",
                               yaxis_title="", margin=dict(t=40, b=0, l=0, r=10))
        st.plotly_chart(fig_bar, use_container_width=True)

    # Target progress
    st.divider()
    st.subheader("🎯 Progress to Target")
    prog_cols = st.columns(3)
    with prog_cols[0]:
        pct5 = min(progress["pct_to_5x"], 100)
        st.markdown(f"**5x Target ({ccy_sym}{TARGET_5X_USD/1e6:.2f}M by 2035)**")
        st.progress(pct5 / 100)
        st.caption(f"{pct5:.1f}% there  |  Gap: {ccy_sym}{progress['gap_5x']:,.0f}  |  Need {progress['cagr_needed_5x']:.1f}% CAGR")
    with prog_cols[1]:
        pct10 = min(progress["pct_to_10x"], 100)
        st.markdown(f"**10x Buffer ({ccy_sym}{TARGET_10X_USD/1e6:.2f}M by 2042)**")
        st.progress(pct10 / 100)
        st.caption(f"{pct10:.1f}% there")
    with prog_cols[2]:
        fig_gauge = go.Figure(go.Indicator(
            mode="gauge+number+delta",
            value=total_mv / 1e6,
            number={"prefix": ccy_sym, "suffix": "M"},
            delta={"reference": TARGET_5X_USD / 1e6, "suffix": "M target"},
            gauge={
                "axis": {"range": [0, TARGET_5X_USD / 1e6]},
                "bar":  {"color": "#2E75B6"},
                "steps": [
                    {"range": [0, TARGET_5X_USD * 0.5 / 1e6], "color": "#1e293b"},
                    {"range": [TARGET_5X_USD * 0.5 / 1e6, TARGET_5X_USD / 1e6], "color": "#0f2d4a"},
                ],
                "threshold": {"line": {"color": "#4ade80", "width": 3},
                              "thickness": 0.75, "value": TARGET_5X_USD / 1e6},
            },
            title={"text": "Portfolio vs 5x Target"},
        ))
        fig_gauge.update_layout(height=220, margin=dict(t=30, b=10, l=20, r=20))
        st.plotly_chart(fig_gauge, use_container_width=True)

    # Top 10 holdings
    st.divider()
    st.subheader("Top 10 Positions")
    top10 = port_df[port_df["mv_usd"].notna()].nlargest(10, "mv_usd")
    total_port = port_df["mv_usd"].sum()
    top10_display = pd.DataFrame({
        "Ticker":    top10["ticker"],
        "Name":      top10["name"],
        "MV (USD)":  top10["mv_usd"].apply(lambda x: f"${x:,.0f}"),
        "% Port":    (top10["mv_usd"] / total_port * 100).apply(lambda x: f"{x:.1f}%"),
        "G/L":       top10["gl_usd"].apply(lambda x: f"${x:+,.0f}" if pd.notna(x) else "—"),
        "G/L %":     top10["gl_pct"].apply(lambda x: f"{x:+.1f}%" if pd.notna(x) else "—"),
        "Barbell":   top10["barbell_class"],
        "Region":    top10["region"],
    })
    st.dataframe(top10_display, use_container_width=True, hide_index=True)

    # ── Portfolio Value Over Time ─────────────────────────────────
    st.divider()
    st.subheader("📈 Portfolio Value Over Time")

    if SHEETS_AVAILABLE:
        _hist = read_portfolio_history()

        if len(_hist) < 2:
            st.info(
                "Portfolio history builds automatically each day you open the app. "
                "Come back tomorrow for your first chart — or click the button below "
                "to save today's snapshot now."
            )
            if st.button("📸 Save today's snapshot", key="snap_now_inline"):
                try:
                    _mv_usd = summary["total_mv"] if report_ccy == "USD" else summary["total_mv"] / fx_rate
                    append_portfolio_snapshot({
                        "total_mv_usd":   _mv_usd,
                        "total_cost_usd": summary["total_cost"] if report_ccy == "USD"
                                          else summary["total_cost"] / fx_rate,
                        "total_gl_usd":   summary["total_gl"] if report_ccy == "USD"
                                          else summary["total_gl"] / fx_rate,
                        "gl_pct":         summary["total_gl_pct"],
                        "hkd_usd_rate":   fx_rate,
                    })
                    read_portfolio_history.clear()
                    st.success("Snapshot saved! Reload the page to see the chart.")
                    st.rerun()
                except Exception as _e:
                    st.error(f"Save failed: {_e}")
        else:
            # Main chart
            _fig_hist = go.Figure()
            _fig_hist.add_trace(go.Scatter(
                x=_hist["date"], y=_hist["total_mv_usd"],
                name="Portfolio Value",
                line=dict(color="#2E75B6", width=2),
                fill="tozeroy", fillcolor="rgba(46,117,182,0.1)",
            ))
            _fig_hist.add_trace(go.Scatter(
                x=_hist["date"], y=_hist["total_cost_usd"],
                name="Cost Basis",
                line=dict(color="#888", width=1, dash="dash"),
            ))
            _fig_hist.add_hline(
                y=13_350_000,
                line_dash="dot", line_color="green",
                annotation_text="5x Target $13.35M",
                annotation_position="bottom right",
            )
            _fig_hist.update_layout(
                title="Portfolio Value vs Cost Basis (USD)",
                xaxis_title="Date", yaxis_title="Value (USD)",
                hovermode="x unified", height=420,
                legend=dict(orientation="h", y=1.05),
            )
            st.plotly_chart(_fig_hist, use_container_width=True)

            # Stats row
            _h_first = _hist.iloc[0]
            _h_last  = _hist.iloc[-1]
            _h_days  = (_h_last["date"] - _h_first["date"]).days
            _h_chg   = _h_last["total_mv_usd"] - _h_first["total_mv_usd"]
            _h_ptt   = _h_last["total_mv_usd"] / 13_350_000 * 100

            _hc1, _hc2, _hc3, _hc4 = st.columns(4)
            _hc1.metric("Tracking since", f"{_h_days} days")
            _hc2.metric("Change since first record", f"${_h_chg:+,.0f}")
            _hc3.metric("Progress to 5x target",    f"{_h_ptt:.1f}%")
            _hc4.metric("Snapshots saved",           len(_hist))

            # Manual snapshot + raw data
            _snap_col, _data_col = st.columns([1, 2])
            with _snap_col:
                if st.button("📸 Save snapshot now", key="snap_manual"):
                    try:
                        _mv_usd2 = (summary["total_mv"] if report_ccy == "USD"
                                    else summary["total_mv"] / fx_rate)
                        append_portfolio_snapshot({
                            "total_mv_usd":   _mv_usd2,
                            "total_cost_usd": summary["total_cost"] if report_ccy == "USD"
                                              else summary["total_cost"] / fx_rate,
                            "total_gl_usd":   summary["total_gl"] if report_ccy == "USD"
                                              else summary["total_gl"] / fx_rate,
                            "gl_pct":         summary["total_gl_pct"],
                            "hkd_usd_rate":   fx_rate,
                        })
                        read_portfolio_history.clear()
                        st.success("Snapshot saved!")
                        st.rerun()
                    except Exception as _e:
                        st.error(f"Save failed: {_e}")

            with _data_col:
                with st.expander("View history data"):
                    _hdisp = _hist.copy()
                    _hdisp["date"]          = _hdisp["date"].dt.strftime("%Y-%m-%d")
                    _hdisp["total_mv_usd"]  = _hdisp["total_mv_usd"].apply(lambda x: f"${x:,.0f}")
                    _hdisp["total_cost_usd"]= _hdisp["total_cost_usd"].apply(lambda x: f"${x:,.0f}")
                    _hdisp["gl_pct"]        = _hdisp["gl_pct"].apply(lambda x: f"{x:.2f}%")
                    st.dataframe(_hdisp[["date","total_mv_usd","total_cost_usd","gl_pct"]],
                                 use_container_width=True, hide_index=True)
    else:
        st.caption("Connect Google Sheets to enable portfolio history tracking.")

    st.divider()
    _pdf3_col, _ = st.columns([1, 3])
    with _pdf3_col:
        if st.button("📄 Generate Portfolio PDF", key="pdf_btn_tab3",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf3 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF",
                _pdf3,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab3",
            )


# ══════════════════════════════════════════════════════════════════
# TAB 4 — ALERTS (Concentration + Compliance)
# ══════════════════════════════════════════════════════════════════
with tab4:
    st.subheader("⚠️ Alerts & Compliance")

    # Concentration
    st.markdown("#### Concentration Alerts  (positions > 5% of portfolio)")
    conc = concentration_alerts(port_df)
    if not conc.empty:
        conc_display = pd.DataFrame({
            "Ticker":       conc["ticker"],
            "Name":         conc["name"],
            "Sector":       conc["sector"],
            "Region":       conc["region"],
            "MV (USD)":     conc["mv_usd"].apply(lambda x: f"${x:,.0f}"),
            "% Portfolio":  conc["pct_of_port"].apply(lambda x: f"{x:.1f}%"),
            "Held at":      conc["brokers"],
        })
        st.dataframe(conc_display, use_container_width=True, hide_index=True)
    else:
        st.success("No single position exceeds 5% of portfolio.")

    # Cross-custodian fragmentation
    st.markdown("#### Cross-Custodian Fragmentation")
    multi = port_df[port_df["brokers"].str.contains(",")].copy()
    multi = multi.sort_values("mv_usd", ascending=False, na_position="last")
    if not multi.empty:
        m_display = pd.DataFrame({
            "Ticker":    multi["ticker"],
            "Name":      multi["name"],
            "MV (USD)":  multi["mv_usd"].apply(lambda x: f"${x:,.0f}" if pd.notna(x) else "—"),
            "Held at":   multi["brokers"],
        })
        st.dataframe(m_display, use_container_width=True, hide_index=True)

    # Compliance
    st.markdown("#### Compliance Check")
    comp = compliance_check(port_df)
    if not comp.empty:
        st.error(f"⛔ {len(comp)} compliance issue(s) found")
        st.dataframe(comp[["ticker", "name", "sector", "flags"]],
                     use_container_width=True, hide_index=True)
    else:
        st.success("✅ No compliance violations detected.")

    # Positions with no price
    st.markdown("#### Positions Missing Live Price")
    no_price = port_df[port_df["live_price"].isna()][["ticker", "name", "sector", "price_error"]]
    if not no_price.empty:
        st.warning(f"{len(no_price)} position(s) have no live price")
        st.dataframe(no_price, use_container_width=True, hide_index=True)
    else:
        st.success("✅ All positions have live prices.")

    st.divider()
    _pdf4_col, _ = st.columns([1, 3])
    with _pdf4_col:
        if st.button("📄 Generate Portfolio PDF", key="pdf_btn_tab4",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf4 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF",
                _pdf4,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab4",
            )


# ══════════════════════════════════════════════════════════════════
# TAB 5 — TRADE ENTRY
# ══════════════════════════════════════════════════════════════════
with tab5:
    st.subheader("✏️ Trade Entry")
    st.caption("Enter trades manually. The app calculates new cost basis and updates holdings automatically.")

    entry_mode = st.radio("Entry mode", ["Single trade", "Upload activity CSV"],
                           horizontal=True)

    if entry_mode == "Single trade":

        # ── Ticker lookup (outside form so it triggers reruns) ────
        @st.cache_data(ttl=3600)
        def _lookup_name(tkr: str) -> str:
            try:
                import yfinance as yf
                info = yf.Ticker(tkr).info
                return info.get("shortName") or info.get("longName") or ""
            except Exception:
                return ""

        # Reset flag must be consumed before the widget renders
        if st.session_state.pop("_reset_te_ticker", False):
            st.session_state.pop("te_ticker", None)

        _te_col1, _te_col2 = st.columns([2, 3])
        with _te_col1:
            _te_ticker = st.text_input(
                "Ticker",
                key="te_ticker",
                placeholder="e.g. MA, MSFT, 0700.HK",
            )
        with _te_col2:
            st.write("")
            st.write("")
            if _te_ticker.strip():
                _te_name = _lookup_name(_te_ticker.strip().upper())
                if _te_name:
                    st.success(f"📌 {_te_name}")
                else:
                    st.warning("Ticker not recognised — double-check before submitting.")

        with st.form("trade_form", clear_on_submit=True):
            c1, c2, c3 = st.columns(3)
            with c1:
                broker = st.selectbox("Broker", BROKERS)
                action = st.radio("Action", ["BUY", "SELL"], horizontal=True)
            with c2:
                ticker = st.text_input(
                    "Ticker (confirm)",
                    value=st.session_state.get("te_ticker", ""),
                    placeholder="e.g. MSFT or 0700.HK",
                )
                ccy    = st.radio("Currency", ["USD", "HKD"], horizontal=True)
            with c3:
                trade_date = st.date_input("Trade date", value=datetime.date.today())
                settle_date= st.date_input("Settle date",
                                            value=datetime.date.today() + datetime.timedelta(days=2))

            c4, c5, c6 = st.columns(3)
            with c4:
                shares = st.number_input("Shares", min_value=0.0, step=1.0, format="%.4f")
            with c5:
                price  = st.number_input("Price (local ccy)", min_value=0.0, step=0.01, format="%.4f")
            with c6:
                comm   = st.number_input("Commission", min_value=0.0, step=0.01, format="%.2f")

            notes  = st.text_input("Notes (optional)")
            gross  = shares * price

            # Preview
            if shares > 0 and price > 0:
                st.info(f"**Preview:** {action} {shares:,.2f} × "
                        f"{'HK$' if ccy=='HKD' else '$'}{price:,.4f} = "
                        f"{'HK$' if ccy=='HKD' else '$'}{gross:,.2f}  "
                        f"(+ {'HK$' if ccy=='HKD' else '$'}{comm:.2f} commission)")

                # Show impact on existing position
                existing = port_df[port_df["ticker"] == ticker.upper().strip()]
                if not existing.empty and action == "BUY":
                    cur_shares = existing["shares"].iloc[0]
                    cur_cost   = existing["cost_local"].iloc[0]
                    new_avg    = calc_new_avg_cost(cur_shares, cur_cost, shares, price)
                    new_total  = cur_shares + shares
                    st.success(f"**Cost basis impact:** "
                               f"Current avg {'HK$' if ccy=='HKD' else '$'}{cur_cost:,.4f} → "
                               f"New avg {'HK$' if ccy=='HKD' else '$'}{new_avg:,.4f}  |  "
                               f"Total shares: {cur_shares:,.2f} → {new_total:,.2f}")

            submitted = st.form_submit_button("✅ Confirm Trade", type="primary")
            if submitted:
                if not ticker.strip():
                    st.error("Please enter a ticker symbol.")
                elif shares <= 0 or price <= 0:
                    st.error("Shares and price must be greater than zero.")
                else:
                    trade = {
                        "trade_date":  str(trade_date),
                        "settle_date": str(settle_date),
                        "broker":      broker,
                        "ticker":      ticker.strip().upper(),
                        "action":      action,
                        "shares":      shares,
                        "price_local": price,
                        "ccy":         ccy,
                        "commission":  comm,
                        "fx_rate":     fx_rate if ccy == "HKD" else 1.0,
                        "gross_usd":   round(gross / fx_rate, 2) if ccy == "HKD" else round(gross, 2),
                        "notes":       notes,
                    }
                    with st.spinner("Saving trade and updating holdings…"):
                        append_trades([trade], source="manual")
                    st.success(f"✅ Trade recorded: {action} {shares:,.2f} {ticker} @ "
                               f"{'HK$' if ccy=='HKD' else '$'}{price:,.4f}")
                    st.session_state["_reset_te_ticker"] = True
                    st.cache_data.clear()
                    st.rerun()

    else:  # Upload CSV
        st.markdown("Upload a transaction CSV from any of the active brokers:")
        broker_up = st.selectbox("Which broker?", ACTIVE_BROKERS)
        uploaded  = st.file_uploader(f"Upload {broker_up} activity CSV",
                                      type=["csv"], key="upload_csv")
        if uploaded:
            from parsers import PARSER_MAP
            try:
                parser = PARSER_MAP[broker_up]
                trades = parser(uploaded)
                st.success(f"✅ Parsed {len(trades)} trades from {broker_up}")

                # Preview table
                preview_df = pd.DataFrame(trades)[
                    ["trade_date", "ticker", "action", "shares", "price_local", "ccy", "commission"]
                ]
                st.dataframe(preview_df, use_container_width=True, hide_index=True)

                if st.button("📥 Import all trades to Google Sheets", type="primary"):
                    with st.spinner("Writing to Google Sheets…"):
                        append_trades(trades, source=f"{broker_up.lower()}_upload")
                    st.success(f"✅ {len(trades)} trades imported. Portfolio recalculated.")
                    st.cache_data.clear()
                    st.rerun()

            except Exception as e:
                st.error(f"Parse error: {e}")
                st.caption("Check that you exported the correct report type. See SETUP.md for format details.")

    # Recent trades log
    st.divider()
    st.markdown("#### Recent Trades Log")
    if SHEETS_AVAILABLE:
        try:
            trades_df = read_trades()
            if not trades_df.empty:
                recent = trades_df.sort_values("trade_date", ascending=False).head(20)
                st.dataframe(
                    recent[["trade_date", "broker", "ticker", "action",
                             "shares", "price_local", "ccy", "commission", "source"]],
                    use_container_width=True, hide_index=True,
                )
            else:
                st.caption("No trades recorded yet.")
        except Exception as e:
            st.caption(f"Could not load trade log: {e}")
    else:
        st.caption("Connect Google Sheets to see trade log (see SETUP.md).")

    st.divider()
    _pdf5_col, _ = st.columns([1, 3])
    with _pdf5_col:
        if st.button("📄 Generate Portfolio PDF", key="pdf_btn_tab5",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf5 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF",
                _pdf5,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab5",
            )


# ══════════════════════════════════════════════════════════════════
# TAB 6 — EXPORT
# ══════════════════════════════════════════════════════════════════
with tab6:
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    st.subheader("📄 Export Portfolio Report")
    st.caption("Downloads a formatted Excel workbook with all tabs. Open in Excel → File → Print → Save as PDF.")

    def build_excel(port_df, summary, fx_rate, report_ccy) -> bytes:
        wb = openpyxl.Workbook()

        # ── Styles ────────────────────────────────────────────────
        hdr_font    = Font(bold=True, color="FFFFFF", size=11)
        hdr_fill    = PatternFill("solid", fgColor="1F3864")
        alt_fill    = PatternFill("solid", fgColor="EEF2F7")
        title_font  = Font(bold=True, size=13)
        border_side = Side(style="thin", color="CCCCCC")
        thin_border = Border(bottom=border_side)
        sym         = "HK$" if report_ccy == "HKD" else "$"

        def style_header_row(ws, row_num, ncols):
            for c in range(1, ncols + 1):
                cell = ws.cell(row=row_num, column=c)
                cell.font = hdr_font
                cell.fill = hdr_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        def style_data_rows(ws, start_row, end_row, ncols):
            for r in range(start_row, end_row + 1):
                fill = alt_fill if r % 2 == 0 else None
                for c in range(1, ncols + 1):
                    cell = ws.cell(row=r, column=c)
                    if fill:
                        cell.fill = fill
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical="center")

        def autofit(ws, min_w=8, max_w=40):
            for col in ws.columns:
                max_len = max((len(str(cell.value or "")) for cell in col), default=0)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, min_w), max_w)

        # ── Sheet 1: Summary ──────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Portfolio Summary"
        ws1.row_dimensions[1].height = 30

        import datetime as _dt
        hkt = _dt.timezone(_dt.timedelta(hours=8))
        now_str = _dt.datetime.now(hkt).strftime("%Y-%m-%d %H:%M HKT")

        summary_data = [
            ["Project Apex 2035 — Portfolio Report", ""],
            [f"Generated", now_str],
            [f"Report Currency", report_ccy],
            [f"FX Rate (USD/HKD)", f"{fx_rate:.4f}"],
            ["", ""],
            ["METRIC", "VALUE"],
            ["Total Market Value", f"{sym}{summary['total_mv']:,.0f}"],
            ["Total Cost Basis", f"{sym}{summary['total_cost']:,.0f}"],
            ["Unrealized G/L", f"{sym}{summary['total_gl']:+,.0f}"],
            ["G/L %", f"{summary['total_gl_pct']:+.2f}%"],
            ["Number of Positions", summary["n_positions"]],
            ["5x Target", f"{sym}{summary['target_5x']:,.0f}"],
            ["% to 5x Target", f"{summary['pct_to_5x']:.1f}%"],
            ["10x Target", f"{sym}{summary['target_10x']:,.0f}"],
            ["% to 10x Target", f"{summary['pct_to_10x']:.1f}%"],
        ]
        for i, row in enumerate(summary_data, 1):
            ws1.cell(i, 1, row[0])
            ws1.cell(i, 2, row[1])
        ws1.cell(1, 1).font = title_font
        style_header_row(ws1, 6, 2)
        style_data_rows(ws1, 7, len(summary_data), 2)
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 24

        # ── Sheet 2: All Positions ────────────────────────────────
        ws2 = wb.create_sheet("All Positions")
        headers = ["Ticker", "Name", "Region", "Sector", "Barbell",
                   "CCY", "Shares", "Avg Cost (Local)", "Live Price",
                   f"MV ({report_ccy})", f"Cost ({report_ccy})",
                   "G/L (USD)", "G/L %", "Held At", "Price Time"]
        for c, h in enumerate(headers, 1):
            ws2.cell(1, c, h)
        style_header_row(ws2, 1, len(headers))
        ws2.row_dimensions[1].height = 28

        df_sorted = port_df.sort_values("mv_usd", ascending=False, na_position="last")
        for r, (_, row) in enumerate(df_sorted.iterrows(), 2):
            mv_r  = row["mv_report"]
            cost_r = row["cost_usd"] * (fx_rate if report_ccy == "HKD" else 1) if row["cost_usd"] else None
            ws2.cell(r, 1,  row["ticker"])
            ws2.cell(r, 2,  row["name"])
            ws2.cell(r, 3,  row["region"])
            ws2.cell(r, 4,  row["sector"])
            ws2.cell(r, 5,  row["barbell_class"])
            ws2.cell(r, 6,  row["ccy"])
            ws2.cell(r, 7,  row["shares"])
            ws2.cell(r, 8,  row["cost_local"])
            ws2.cell(r, 9,  row["live_price"])
            ws2.cell(r, 10, mv_r)
            ws2.cell(r, 11, cost_r)
            ws2.cell(r, 12, row["gl_usd"])
            ws2.cell(r, 13, f"{row['gl_pct']:+.1f}%" if row["gl_pct"] is not None and str(row["gl_pct"]) != "nan" else "—")
            ws2.cell(r, 14, row["brokers"])
            ws2.cell(r, 15, row["price_ts"])
            # Number formats
            for col in [7, 8, 9, 10, 11, 12]:
                cell = ws2.cell(r, col)
                if cell.value is not None:
                    cell.number_format = '#,##0.00'
        style_data_rows(ws2, 2, len(df_sorted) + 1, len(headers))
        autofit(ws2)

        # ── Sheet 3: By Broker ────────────────────────────────────
        ws3 = wb.create_sheet("By Broker")
        broker_headers = ["Broker", "Ticker", "Name", "Shares",
                          "Avg Cost (Local)", "Live Price", f"MV ({report_ccy})"]
        for c, h in enumerate(broker_headers, 1):
            ws3.cell(1, c, h)
        style_header_row(ws3, 1, len(broker_headers))
        ws3.row_dimensions[1].height = 28

        r = 2
        for broker in BROKERS:
            for _, row in df_sorted.iterrows():
                b_detail = next((b for b in row["brokers_list"] if b["broker"] == broker), None)
                if b_detail is None:
                    continue
                sh = b_detail["shares"]
                lp = row["live_price"]
                mv_here = None
                if lp and sh:
                    mv_local_h = sh * lp
                    mv_usd_h   = mv_local_h / fx_rate if row["ccy"] == "HKD" else mv_local_h
                    mv_here    = mv_usd_h * fx_rate if report_ccy == "HKD" else mv_usd_h
                ws3.cell(r, 1, broker)
                ws3.cell(r, 2, row["ticker"])
                ws3.cell(r, 3, row["name"])
                ws3.cell(r, 4, sh)
                ws3.cell(r, 5, b_detail["avg_cost_local"])
                ws3.cell(r, 6, lp)
                ws3.cell(r, 7, mv_here)
                for col in [4, 5, 6, 7]:
                    cell = ws3.cell(r, col)
                    if cell.value is not None:
                        cell.number_format = '#,##0.00'
                r += 1
        style_data_rows(ws3, 2, r - 1, len(broker_headers))
        autofit(ws3)

        # ── Sheet 4: By Sector ────────────────────────────────────
        ws4 = wb.create_sheet("By Sector")
        alloc_h = ["Group", "Sector", "Ticker", "Name", f"MV ({report_ccy})", "% Portfolio"]
        for c, h in enumerate(alloc_h, 1):
            ws4.cell(1, c, h)
        style_header_row(ws4, 1, len(alloc_h))
        ws4.row_dimensions[1].height = 28

        total_mv_usd = port_df["mv_usd"].sum()
        df_sec = df_sorted[df_sorted["mv_usd"].notna()].copy()
        r = 2
        for barbell in ["CORE", "TACTICAL", "SPECULATIVE"]:
            grp = df_sec[df_sec["barbell_class"] == barbell]
            for _, row in grp.iterrows():
                mv_r = row["mv_report"]
                pct  = row["mv_usd"] / total_mv_usd * 100 if total_mv_usd else 0
                ws4.cell(r, 1, barbell)
                ws4.cell(r, 2, row["sector"])
                ws4.cell(r, 3, row["ticker"])
                ws4.cell(r, 4, row["name"])
                ws4.cell(r, 5, mv_r)
                ws4.cell(r, 6, f"{pct:.2f}%")
                if ws4.cell(r, 5).value is not None:
                    ws4.cell(r, 5).number_format = '#,##0'
                r += 1
        style_data_rows(ws4, 2, r - 1, len(alloc_h))
        autofit(ws4)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    col_btn, col_info = st.columns([1, 2])
    with col_btn:
        if st.button("📥 Generate Excel Report", type="primary"):
            xlsx_bytes = build_excel(port_df, summary, fx_rate, report_ccy)
            st.download_button(
                label="⬇️ Download Excel (.xlsx)",
                data=xlsx_bytes,
                file_name=f"Apex2035_Portfolio_{datetime.date.today()}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
    with col_info:
        st.markdown("""
        **What's included:**
        - **Sheet 1:** Portfolio Summary (totals, targets, FX)
        - **Sheet 2:** All Positions — sorted by MV, all metrics
        - **Sheet 3:** By Broker — every position per broker
        - **Sheet 4:** By Sector/Barbell — allocation breakdown

        **To save as PDF:** Open in Excel → File → Print → Microsoft Print to PDF
        """)

    st.divider()
    st.subheader("📄 PDF Report")
    st.caption("Text-based PDF — selectable, copyable, ready for Claude/Gemini.")
    pdf6_col, pdf6_info = st.columns([1, 2])
    with pdf6_col:
        if st.button("📥 Generate PDF Report", key="pdf_btn_tab6", type="primary",
                     use_container_width=True):
            with st.spinner("Generating PDF…"):
                _pdf6 = export_portfolio_pdf(
                    port_df, summary, fx_rate, report_ccy,
                    concentration_alerts(port_df),
                    compliance_check(port_df),
                )
            st.download_button(
                "⬇️ Download PDF (.pdf)",
                _pdf6,
                f"apex2035_portfolio_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                "application/pdf",
                key="dl_pdf_tab6",
            )
    with pdf6_info:
        st.markdown("""
        **PDF includes (4 pages):**
        - **Page 1:** Executive summary — totals, top 5, alerts count
        - **Page 2:** Allocation by barbell / region / sector
        - **Page 3:** Full holdings sorted by market value
        - **Page 4:** Concentration alerts, compliance, missing prices
        """)


# ══════════════════════════════════════════════════════════════════
# TAB 7 — STOCK ANALYZER
# ══════════════════════════════════════════════════════════════════
with tab7:
    # Session state for persistent results across reruns
    if "az_result"  not in st.session_state: st.session_state.az_result  = None
    if "az_ticker"  not in st.session_state: st.session_state.az_ticker  = ""
    if "az_fv"      not in st.session_state: st.session_state.az_fv      = None
    if "lseg_calls" not in st.session_state: st.session_state.lseg_calls = 0

    st.subheader("📈 Stock Analyzer — 10-Pillar Framework")
    st.caption("Analyze any ticker: MSFT, AAPL (US) or 0700.HK, 9988.HK (HK stocks).")

    # ── Input + buttons ───────────────────────────────────────────
    _a1, _a2, _a3 = st.columns([3, 1, 1])
    with _a1:
        _aticker = st.text_input("Ticker", placeholder="e.g. MSFT, 0700.HK",
                                  key="analyzer_ticker_input")
    with _a2:
        st.write(""); st.write("")
        _analyze_btn = st.button("🔍 Analyze", type="primary", key="analyze_btn")
    with _a3:
        st.write(""); st.write("")
        _lseg_btn = st.button(
            "🔬 + LSEG",
            key="lseg_enhance_btn",
            disabled=not _lseg_ok,
            help=(
                "Enhance with LSEG data from Refinitiv Workspace."
                if _lseg_ok
                else "Requires Refinitiv Workspace open on this PC. "
                     "Not available on the cloud version."
            ),
        )
    if not _lseg_ok:
        st.caption("💡 **+ LSEG** only works locally with Workspace open.")

    # ── Handle button clicks ──────────────────────────────────────
    if _analyze_btn and _aticker.strip():
        _tc = _aticker.strip().upper()
        calculate_pillars.clear()
        with st.spinner(f"Analyzing {_tc} via yfinance… (10–20 s)"):
            _r = calculate_pillars(_tc, False)
        st.session_state.az_result = _r
        st.session_state.az_ticker = _tc
        st.session_state.az_fv     = None   # reset fair value on new ticker
    elif _analyze_btn:
        st.warning("Please enter a ticker symbol.")

    if _lseg_btn:
        _tc = (_aticker.strip().upper() or st.session_state.az_ticker)
        if not _tc:
            st.warning("Enter a ticker first, then click Analyze, then + LSEG.")
        else:
            st.warning("⚠️ Using corporate LSEG account — use sparingly.")
            calculate_pillars.clear()
            with st.spinner(f"Enhancing {_tc} with LSEG data…"):
                _r = calculate_pillars(_tc, True)
            st.session_state.az_result = _r
            st.session_state.az_ticker = _tc
            st.session_state.lseg_calls += 1
            st.success(f"✓ LSEG data loaded  |  Session total: {st.session_state.lseg_calls} call(s)")

    # ── Display (from session state — persists across reruns) ─────
    _result = st.session_state.az_result
    _ticker_clean = st.session_state.az_ticker

    if _result:
        if _result.get("error"):
            st.error(f"❌ {_result['error']}")
            st.caption("Tips: US stocks: MSFT | HK stocks need .HK suffix: 0700.HK | ETFs show N/A on most pillars.")
        else:
            # ── Company header ────────────────────────────────────
            _info  = _result["company_info"]
            _is_hk = _ticker_clean.endswith(".HK")
            _px_sym = "HK$" if _is_hk else "$"

            _h1, _h2, _h3, _h4 = st.columns(4)
            _h1.metric("Company",    str(_info.get("name", "N/A"))[:30])
            _h2.metric("Sector",     str(_info.get("sector") or "N/A")[:25])
            _mc = _info.get("market_cap")
            _mc_str = (f"${_mc/1e12:.1f}T" if _mc and _mc >= 1e12
                       else f"${_mc/1e9:.1f}B" if _mc and _mc >= 1e9
                       else f"${_mc/1e6:.0f}M" if _mc else "N/A")
            _h3.metric("Market Cap", _mc_str)
            _px = _info.get("price")
            _h4.metric("Price", f"{_px_sym}{_px:,.2f}" if _px else "N/A")

            _w52h = _info.get("week_52_high")
            _w52l = _info.get("week_52_low")
            if _w52h and _w52l:
                st.caption(f"52-week: {_px_sym}{_w52l:,.2f} – {_px_sym}{_w52h:,.2f}"
                           f"  |  Currency: {_info.get('currency', 'USD')}")
            st.divider()

            # ── Verdict ───────────────────────────────────────────
            _score   = _result["score"]
            _verdict = _result["verdict"]
            _vstyle  = {
                "CHEAP":     ("#d4edda", "#1a5c2a", "🟢 CHEAP"),
                "FAIR":      ("#fff3cd", "#856404", "🟡 FAIR"),
                "EXPENSIVE": ("#f8d7da", "#721c24", "🔴 EXPENSIVE"),
            }
            _vbg, _vtc, _vlabel = _vstyle.get(_verdict, ("#f5f5f5", "#333", f"⚫ {_verdict}"))
            _v1, _v2 = st.columns([3, 1])
            with _v1:
                st.markdown(
                    f'<div style="background:{_vbg};border-radius:8px;padding:12px 18px;">'
                    f'<span style="font-size:1.4em;font-weight:700;color:{_vtc}">{_vlabel}</span>'
                    f'</div>', unsafe_allow_html=True)
            _v2.metric("Pillars Passing", f"{_score} / 10")
            st.divider()

            # ── 10-Pillar table ───────────────────────────────────
            _PMETA = {
                1:  ("↓ cheaper vs history",  "🟢 below 5yr avg  🟡 within ±10%  🔴 >10% above avg"),
                2:  ("↑ higher is better",     "🟢 >15%  🟡 10–15%  🔴 <10%"),
                3:  ("↑ higher is better",     "🟢 >10% CAGR  🟡 5–10%  🔴 <5%"),
                4:  ("↑ higher is better",     "🟢 >10% CAGR  🟡 5–10%  🔴 <5%"),
                5:  ("↓ buybacks are better",  "🟢 shrinking  🟡 flat ±2%  🔴 growing (dilution)"),
                6:  ("↓ less debt is better",  "🟢 <2×  🟡 2–4×  🔴 >4×"),
                7:  ("↑ higher is better",     "🟢 >10% CAGR  🟡 5–10%  🔴 <5%"),
                8:  ("↓ cheaper is better",    "🟢 <20×  🟡 20–30×  🔴 >30×"),
                9:  ("↑ expanding is better",  "🟢 >+2pp  🟡 stable ±2pp  🔴 shrinking"),
                10: ("↑ beat the index",       "🟢 outperform >5%  🟡 within ±5%  🔴 trail >5%"),
            }
            _RBADGE = {
                "GREEN":  '<span style="background:#d4edda;color:#155724;padding:3px 10px;border-radius:12px;font-weight:700;font-size:12px">✅ PASS</span>',
                "YELLOW": '<span style="background:#fff3cd;color:#856404;padding:3px 10px;border-radius:12px;font-weight:700;font-size:12px">⚠️ FAIR</span>',
                "RED":    '<span style="background:#f8d7da;color:#721c24;padding:3px 10px;border-radius:12px;font-weight:700;font-size:12px">❌ FAIL</span>',
                "NA":     '<span style="background:#e9ecef;color:#6c757d;padding:3px 10px;border-radius:12px;font-size:12px">— N/A</span>',
            }
            _rows = ""
            for _i, _p in enumerate(_result.get("pillars", [])):
                _n = _p["number"]
                _md, _mt = _PMETA.get(_n, ("", ""))
                _badge = _RBADGE.get(_p["rating"], _RBADGE["NA"])
                _bg = "#fafafa" if _i % 2 == 0 else "#ffffff"
                _rows += (
                    f'<tr style="background:{_bg}">'
                    f'<td style="color:#999;font-size:11px;padding:8px 6px">{_n}</td>'
                    f'<td style="padding:8px 6px;font-weight:600">{_p["name"]}</td>'
                    f'<td style="padding:8px 6px;font-family:monospace;font-size:13px">{_p["value"]}</td>'
                    f'<td style="padding:8px 6px;text-align:center">{_badge}</td>'
                    f'<td style="padding:8px 6px;color:#555;font-size:12px">{_md}</td>'
                    f'<td style="padding:8px 6px;color:#666;font-size:11px">{_mt}</td>'
                    f'</tr>'
                )
            st.markdown(
                '<table style="width:100%;border-collapse:collapse;font-size:13px">'
                '<thead><tr style="background:#1F3864;color:white">'
                '<th style="padding:8px 6px;width:3%">#</th>'
                '<th style="padding:8px 6px;width:17%">Pillar</th>'
                '<th style="padding:8px 6px;width:20%">Value</th>'
                '<th style="padding:8px 6px;width:9%;text-align:center">Score</th>'
                '<th style="padding:8px 6px;width:14%">Direction</th>'
                '<th style="padding:8px 6px;width:37%">Thresholds</th>'
                f'</tr></thead><tbody>{_rows}</tbody></table>',
                unsafe_allow_html=True,
            )
            st.divider()

            # ── Fair Value Calculator ─────────────────────────────
            st.subheader("💰 Fair Value Estimator")
            st.caption("Paul Gabrail / Everything Money DCF methodology. Adjust to match your thesis.")

            # ── Reference data panel ──────────────────────────────
            _hist_data = _result.get("historical", [])
            if _hist_data:
                with st.expander("📊 Actual company data — use as reference for your sliders",
                                  expanded=True):
                    # Compute actual metrics from historical data
                    def _avg(lst):
                        lst = [x for x in lst if x is not None]
                        return sum(lst) / len(lst) if lst else None

                    _rev_vals  = [h.get("revenue")    for h in _hist_data]
                    _ni_vals   = [h.get("net_income")  for h in _hist_data]
                    _fcf_vals  = [h.get("fcf")         for h in _hist_data]
                    _gm_vals   = [h.get("gross_margin_pct") for h in _hist_data]

                    # Revenue growth YoY
                    _rev_g_yoy = []
                    for _i in range(len(_rev_vals) - 1):
                        _r0 = _rev_vals[_i + 1]  # older
                        _r1 = _rev_vals[_i]       # newer
                        if _r0 and _r1 and _r0 > 0:
                            _rev_g_yoy.append((_r1 / _r0 - 1) * 100)

                    # Margins from latest and average
                    _fcf_margins = [
                        f / r * 100 for f, r in zip(_fcf_vals, _rev_vals)
                        if f and r and r > 0
                    ]
                    _ni_margins = [
                        n / r * 100 for n, r in zip(_ni_vals, _rev_vals)
                        if n and r and r > 0
                    ]

                    # Current P/E from Pillar 1 value string
                    _p1_val = _result["pillars"][0]["value"] if _result.get("pillars") else ""
                    _cur_pe_str = _p1_val.split("x")[0] if "x" in _p1_val else "N/A"
                    try:
                        _cur_pe = float(_cur_pe_str)
                    except Exception:
                        _cur_pe = None

                    def _fmt(v, suffix="%", decimals=1):
                        return f"{v:.{decimals}f}{suffix}" if v is not None else "N/A"

                    _ref_cols = st.columns(5)
                    _ref_cols[0].metric(
                        "Rev Growth",
                        _fmt(_rev_g_yoy[0] if _rev_g_yoy else None),
                        f"avg {_fmt(_avg(_rev_g_yoy))} / {len(_rev_g_yoy)}yr",
                        help="Most recent year-on-year revenue growth"
                    )
                    _ref_cols[1].metric(
                        "FCF Margin",
                        _fmt(_fcf_margins[0] if _fcf_margins else None),
                        f"avg {_fmt(_avg(_fcf_margins))} / {len(_fcf_margins)}yr",
                        help="Free cash flow as % of revenue"
                    )
                    _ref_cols[2].metric(
                        "Net Margin",
                        _fmt(_ni_margins[0] if _ni_margins else None),
                        f"avg {_fmt(_avg(_ni_margins))} / {len(_ni_margins)}yr",
                        help="Net income as % of revenue"
                    )
                    _ref_cols[3].metric(
                        "Gross Margin",
                        _fmt(_gm_vals[0] if _gm_vals else None),
                        f"avg {_fmt(_avg(_gm_vals))} / {len(_gm_vals)}yr",
                        help="Gross profit as % of revenue"
                    )
                    _ref_cols[4].metric(
                        "Current P/E",
                        f"{_cur_pe:.1f}×" if _cur_pe else "N/A",
                        help="Trailing P/E — use as anchor for terminal P/E"
                    )
                    st.caption(
                        "💡 Use these as starting points. "
                        "Your sliders should reflect your *expectations*, not just history. "
                        "For terminal P/E: quality compounders typically trade at 20–35×."
                    )

            with st.expander("⚙️ Set Your Assumptions", expanded=True):
                _fv1, _fv2, _fv3 = st.columns(3)
                with _fv1:
                    _fv_rev_g  = st.slider("Revenue Growth (%/yr)", -5, 50, 10,
                                            help="Expected annual revenue growth next 5 years")
                    _fv_pm     = st.slider("Target Profit Margin (%)", 0, 50, 25,
                                            help="Expected net profit margin at exit")
                with _fv2:
                    _fv_fcf_m  = st.slider("FCF Margin (%)", 0, 50, 20,
                                            help="Free cash flow as % of revenue")
                    _fv_rr     = st.slider("Required Return (%)", 5, 30, 10,
                                            help="Your hurdle rate. Higher = more conservative. "
                                                 "10% is typical for equities.")
                with _fv3:
                    _fv_tpe    = st.slider("Terminal P/E Multiple", 5, 50, 25,
                                            help="P/E you'd exit at in year N. "
                                                 "25x is reasonable for quality compounders.")
                    _fv_yrs    = st.radio("Projection Years", [3, 5, 10],
                                           index=1, horizontal=True)

            if st.button("💰 Calculate Fair Value", type="primary", key="fv_btn"):
                from core.engine import calculate_fair_value
                with st.spinner("Calculating…"):
                    _fv = calculate_fair_value(
                        _ticker_clean, _fv_rev_g, _fv_pm,
                        _fv_fcf_m, _fv_rr, _fv_tpe, _fv_yrs,
                    )
                st.session_state.az_fv = _fv

            _fv = st.session_state.az_fv
            if _fv:
                if "error" in _fv:
                    st.error(f"Could not calculate: {_fv['error']}")
                else:
                    _fv_sym = "HK$" if _ticker_clean.endswith(".HK") else "$"
                    if _fv.get("cross_currency"):
                        st.warning(
                            f"⚠️ Note: This company reports financials in "
                            f"**{_fv.get('fin_currency')}** but trades in "
                            f"**{_fv.get('currency')}**. "
                            f"Fair value is calculated in reporting currency — "
                            f"treat as approximate."
                        )
                    _fc1, _fc2, _fc3, _fc4 = st.columns(4)
                    _fc1.metric("Bear Case",  f"{_fv_sym}{_fv['fair_value_bear']:,.2f}")
                    _fc2.metric("Base Case",  f"{_fv_sym}{_fv['fair_value_base']:,.2f}",
                                delta=f"{_fv['upside_pct']:+.1f}% vs current")
                    _fc3.metric("Bull Case",  f"{_fv_sym}{_fv['fair_value_bull']:,.2f}")
                    _fc4.metric("Margin of Safety", f"{_fv['margin_of_safety']:+.1f}%",
                                help=">20% is a good entry margin of safety")

                    _fvv = _fv["verdict"]
                    _fv_color = {"UNDERVALUED": "green", "FAIRLY VALUED": "orange",
                                  "OVERVALUED": "red"}.get(_fvv, "gray")
                    st.markdown(f"### :{_fv_color}[{_fvv}]")
                    st.caption(
                        f"Current: {_fv_sym}{_fv['current_price']:,.2f}  |  "
                        f"Bear: {_fv_sym}{_fv['fair_value_bear']:,.2f}  |  "
                        f"Base: {_fv_sym}{_fv['fair_value_base']:,.2f}  |  "
                        f"Bull: {_fv_sym}{_fv['fair_value_bull']:,.2f}"
                    )
                    st.info("💡 These are YOUR assumptions, not analyst consensus. "
                            "Higher required return = more conservative = larger margin of safety.")

            st.divider()

            # ── Watchlist + PDF ───────────────────────────────────
            _wl_col, _pdf_col = st.columns(2)
            with _wl_col:
                if SHEETS_AVAILABLE:
                    if st.button("📌 Add to Watchlist", key="add_watchlist_btn"):
                        from core.sheets import append_watchlist
                        _wl_res = append_watchlist(
                            _ticker_clean, _info.get("price"), _score, _verdict)
                        if _wl_res is True:
                            st.success(f"✅ {_ticker_clean} added to Watchlist!")
                        else:
                            st.error(f"Watchlist error: {_wl_res}")
                else:
                    st.caption("⚠️ Connect Google Sheets to use watchlist.")

            with _pdf_col:
                if st.button("📄 Generate Stock PDF", key="stock_pdf_btn"):
                    with st.spinner("Generating PDF…"):
                        _spdf = export_stock_pdf(
                            _ticker_clean, _result, st.session_state.az_fv)
                    st.download_button(
                        "⬇️ Download Stock PDF", _spdf,
                        f"apex2035_stock_{_ticker_clean}_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                        "application/pdf", key="dl_stock_pdf")


# ══════════════════════════════════════════════════════════════════
# TAB 8 — CONVICTION TRACKER
# ══════════════════════════════════════════════════════════════════
with tab8:
    st.subheader("🎯 Conviction Tracker — Investment Discipline Engine")
    st.caption(
        "Record your investment thesis before every trade. "
        "Review quarterly and grade yourself honestly."
    )

    if not SHEETS_AVAILABLE:
        st.warning(
            "⚠️ Google Sheets not connected. "
            "Conviction data will not be saved. See SETUP.md."
        )

    # ── SECTION 1: New Conviction Entry ──────────────────────────
    st.markdown("### 📝 Record New Conviction")
    with st.form("conviction_form", clear_on_submit=True):
        _c1, _c2 = st.columns(2)
        with _c1:
            _cv_ticker = st.text_input("Ticker", placeholder="e.g. MSFT or 0700.HK")
            _cv_action = st.selectbox("Action",
                ["NEW POSITION", "ADD TO EXISTING", "REDUCE", "EXIT"])
            _cv_price  = st.number_input("Entry Price", min_value=0.0,
                                          step=0.01, format="%.4f")
            _cv_size   = st.number_input("Position Size (USD)", min_value=0.0,
                                          step=1000.0, format="%.0f")
            _cv_max    = st.number_input("Max Size Cap (USD)",  min_value=0.0,
                                          step=1000.0, format="%.0f",
                                          help="Maximum I will ever put in this position")
        with _c2:
            _cv_horizon = st.slider("Time Horizon (months)", 1, 36, 12)
            _cv_false   = st.number_input(
                "Falsification Price",
                min_value=0.0, step=0.01, format="%.4f",
                help="I am wrong if price drops below this level",
            )
            _cv_opp = st.text_input(
                "Opportunity Cost",
                placeholder="vs buying VOO instead",
            )

        _cv_thesis = st.text_area(
            "Investment Thesis (3 sentences max)",
            max_chars=400,
            placeholder="Why this works, why now, why better than alternatives",
        )
        _cv_bull = st.text_area("Bull Case", max_chars=200,
                                 placeholder="What makes this work")
        _cv_bear = st.text_area("Bear Case — what proves me wrong", max_chars=200)

        _cv_submit = st.form_submit_button("✅ Record Conviction", type="primary")

    if _cv_submit:
        if not _cv_ticker.strip():
            st.error("Please enter a ticker symbol.")
        elif _cv_price <= 0:
            st.error("Entry price must be greater than zero.")
        elif not _cv_thesis.strip():
            st.error("Investment thesis is required.")
        else:
            # Fetch company name from yfinance
            _cv_name = _cv_ticker.strip().upper()
            try:
                import yfinance as _yf
                _cv_info = _yf.Ticker(_cv_name).info
                _cv_name = (_cv_info.get("longName") or
                             _cv_info.get("shortName") or _cv_name)
            except Exception:
                pass

            _cv_row = {
                "ticker":             _cv_ticker.strip().upper(),
                "name":               _cv_name,
                "action":             _cv_action,
                "entry_price":        _cv_price,
                "position_size_usd":  _cv_size,
                "max_size_cap_usd":   _cv_max,
                "thesis":             _cv_thesis.strip(),
                "bull_case":          _cv_bull.strip(),
                "bear_case":          _cv_bear.strip(),
                "falsification_price": _cv_false,
                "time_horizon_months": _cv_horizon,
                "opportunity_cost":   _cv_opp.strip(),
                "status":             "ACTIVE",
            }

            if SHEETS_AVAILABLE:
                from core.sheets import append_conviction
                _cv_res = append_conviction(_cv_row)
                if _cv_res is True:
                    st.success(
                        f"✅ Conviction recorded: {_cv_ticker.strip().upper()} — "
                        f"{_cv_action} @ ${_cv_price:,.2f}  |  "
                        f"Horizon: {_cv_horizon} months  |  "
                        f"Falsification: ${_cv_false:,.2f}"
                    )
                else:
                    st.error(f"Failed to save: {_cv_res}")
            else:
                st.info("✅ Conviction validated. Connect Google Sheets to persist.")

    st.divider()

    # ── SECTION 2: Active Convictions ────────────────────────────
    st.markdown("### 📊 Active Convictions")

    if SHEETS_AVAILABLE:
        from core.sheets import read_convictions, update_conviction
        _all_cv = read_convictions()

        if _all_cv.empty:
            st.info("No convictions recorded yet. Add your first one above.")
        else:
            _active_cv = _all_cv[_all_cv["status"] == "ACTIVE"].copy() \
                if "status" in _all_cv.columns else _all_cv.copy()

            if _active_cv.empty:
                st.info("No active convictions. All positions have been reviewed.")
            else:
                # Fetch current prices for P&L
                _cv_tickers_list = _active_cv["ticker"].dropna().unique().tolist()
                with st.spinner("Fetching current prices…"):
                    _cv_prices = {}
                    for _cvt in _cv_tickers_list:
                        try:
                            import yfinance as _yf2
                            _cvtk = _yf2.Ticker(_cvt)
                            _cvp  = (_cvtk.info.get("currentPrice") or
                                      _cvtk.info.get("regularMarketPrice"))
                            _cv_prices[_cvt] = float(_cvp) if _cvp else None
                        except Exception:
                            _cv_prices[_cvt] = None

                # Build display table
                _cv_rows = []
                for _, _r in _active_cv.iterrows():
                    _t  = str(_r.get("ticker", ""))
                    _ep = _r.get("entry_price")
                    _cp = _cv_prices.get(_t)
                    _pnl_pct = (((_cp - float(_ep)) / float(_ep)) * 100
                                 if (_cp and _ep and float(_ep) > 0) else None)
                    _ed = str(_r.get("entry_date", ""))
                    try:
                        from datetime import date as _date_cls
                        _days = (datetime.date.today() -
                                  datetime.date.fromisoformat(_ed)).days
                    except Exception:
                        _days = 0

                    _fp  = _r.get("falsification_price")
                    _at_false = (_cp and _fp and float(_fp) > 0 and
                                  _cp < float(_fp))

                    _thesis_trunc = str(_r.get("thesis", ""))[:50]
                    _pnl_str = (f"{_pnl_pct:+.1f}%" if _pnl_pct is not None else "—")
                    _cp_str  = (f"${_cp:,.2f}" if _cp else "—")
                    _ep_str  = (f"${float(_ep):,.2f}" if _ep else "—")

                    _cv_rows.append({
                        "Ticker":    _t,
                        "Entry Date":_ed,
                        "Entry $":   _ep_str,
                        "Current $": _cp_str,
                        "P&L %":     _pnl_str,
                        "Days Held": str(_days),
                        "Thesis":    _thesis_trunc + ("…" if len(str(_r.get("thesis",""))) > 50 else ""),
                        "False. $":  (f"${float(_fp):,.2f}" if _fp else "—"),
                        "Alert":     "⚠️ FALSIFICATION" if _at_false else "",
                        "conv_id":   str(_r.get("conviction_id", "")),
                    })

                _cv_display = pd.DataFrame(_cv_rows).drop(columns=["conv_id"])
                st.dataframe(_cv_display, use_container_width=True,
                              hide_index=True)

                # Review forms per conviction (expander per row)
                st.markdown("#### Review a Conviction")
                _cv_options = [
                    f"{r['Ticker']} — {r['Entry Date']} ({r['P&L %']})"
                    for r in _cv_rows
                ]
                _selected_idx = st.selectbox(
                    "Select conviction to review",
                    range(len(_cv_rows)),
                    format_func=lambda i: _cv_options[i],
                    key="cv_review_select",
                )
                if _cv_rows:
                    _sel = _cv_rows[_selected_idx]
                    with st.form("review_form", clear_on_submit=True):
                        _rv_c1, _rv_c2 = st.columns(2)
                        with _rv_c1:
                            _rv_grade  = st.selectbox("Grade",
                                ["A — Thesis played out perfectly",
                                 "B — Mostly correct, minor miss",
                                 "C — Mixed, thesis partly wrong",
                                 "D — Thesis largely wrong",
                                 "F — Completely wrong"])
                            _rv_status = st.selectbox("Update Status",
                                ["ACTIVE", "REVIEWING", "CLOSED"])
                        with _rv_c2:
                            _rv_date = st.date_input("Review Date",
                                                      value=datetime.date.today())
                        _rv_notes = st.text_area("Outcome Notes",
                            placeholder="What happened? What did you miss? What will you do differently?",
                            max_chars=400)
                        _rv_submit = st.form_submit_button(
                            "💾 Save Review", type="primary")

                    if _rv_submit:
                        _rv_grade_letter = _rv_grade[0]  # "A", "B", etc.
                        _rv_res = update_conviction(_sel["conv_id"], {
                            "grade":         _rv_grade_letter,
                            "status":        _rv_status,
                            "review_date":   str(_rv_date),
                            "outcome_notes": _rv_notes.strip(),
                        })
                        if _rv_res is True:
                            st.success("✅ Review saved.")
                            st.cache_data.clear()
                        else:
                            st.error(f"Save failed: {_rv_res}")

        # PDF export
        st.divider()
        _cv_pdf_col, _ = st.columns([1, 2])
        with _cv_pdf_col:
            if st.button("📄 Export Conviction Log PDF", key="cv_pdf_btn",
                         use_container_width=True):
                if not _all_cv.empty:
                    with st.spinner("Generating PDF…"):
                        # Compute basic stats
                        _cl = _all_cv[_all_cv.get("status", pd.Series()) == "CLOSED"] \
                            if "status" in _all_cv.columns else pd.DataFrame()
                        _graded = _cl[_cl["grade"].isin(["A","B","C","D","F"])] \
                            if "grade" in _cl.columns else pd.DataFrame()
                        _wins   = len(_cl[_cl["grade"].isin(["A","B"])]) \
                            if "grade" in _cl.columns else 0
                        _stats = {
                            "total_decisions": len(_all_cv),
                            "win_rate_pct":    (_wins / len(_graded) * 100
                                                if len(_graded) > 0 else 0),
                            "avg_hold_days":   0,
                            "avg_pnl_pct":     0,
                        }
                        _cv_pdf = export_conviction_pdf(_all_cv, _stats)
                    st.download_button(
                        "⬇️ Download Conviction Log",
                        _cv_pdf,
                        f"apex2035_conviction_log_{datetime.date.today().strftime('%Y%m%d')}.pdf",
                        "application/pdf",
                        key="dl_cv_pdf",
                    )
                else:
                    st.info("No conviction data to export yet.")
    else:
        st.caption("Connect Google Sheets to use the Conviction Tracker.")

    st.divider()

    # ── SECTION 3: Statistics ─────────────────────────────────────
    st.markdown("### 📈 Track Record")
    if SHEETS_AVAILABLE:
        try:
            _stats_cv = read_convictions() if not _all_cv.empty else pd.DataFrame()
        except Exception:
            _stats_cv = pd.DataFrame()

        if not _stats_cv.empty and "status" in _stats_cv.columns:
            _closed_cv = _stats_cv[_stats_cv["status"] == "CLOSED"]
            _graded_cv = _closed_cv[_closed_cv["grade"].isin(["A","B","C","D","F"])] \
                if "grade" in _closed_cv.columns else pd.DataFrame()
            _wins_cv   = len(_closed_cv[_closed_cv["grade"].isin(["A","B"])]) \
                if "grade" in _closed_cv.columns else 0

            _s1, _s2, _s3, _s4 = st.columns(4)
            _s1.metric("Total Decisions", len(_stats_cv))
            _s2.metric("Win Rate (A+B)",
                       f"{_wins_cv/len(_graded_cv)*100:.0f}%"
                       if len(_graded_cv) > 0 else "—")
            _s3.metric("Active", len(_stats_cv[_stats_cv["status"] == "ACTIVE"]))
            _s4.metric("Closed", len(_closed_cv))

            if not _graded_cv.empty:
                st.markdown("**Grade Distribution**")
                _gd = _graded_cv["grade"].value_counts().reset_index()
                _gd.columns = ["Grade", "Count"]
                st.dataframe(_gd, use_container_width=False, hide_index=True)
        else:
            st.info("Record and close convictions to see your track record.")
    else:
        st.caption("Connect Google Sheets to see statistics.")


# ══════════════════════════════════════════════════════════════════
# TAB 9 — FLOW MONITOR
# ══════════════════════════════════════════════════════════════════
with tab9:
    from tab_flow_monitor import render_flow_monitor
    render_flow_monitor()
